"""Excel 测试用例批量测试脚本 — 去重 + 全字段输出。

读取 问题.xlsx 的"测试用例"工作表，去重后对每个问题执行 extract → dispatch，
输出包含每个字段具体值的详细测试报告到 requirements/测试报告.md。
"""

import json
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import openpyxl

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from src.pages.infrastructure.llm.llm_gateway import LlmConfig, LlmGateway
from src.pages.request_lifecycle.dispatch_execution import ConcurrentTaskDispatcher
from src.pages.request_lifecycle.unified_intent_extractor import UnifiedIntentExtractor

LLM_CONFIG = LlmConfig(
    api_key="YOUR_DEEPSEEK_API_KEY",
    base_url="https://api.deepseek.com/v1",
    model="deepseek-chat",
    timeout_seconds=90,
)

EXCEL_PATH = Path(r"C:\Users\29029\Downloads\问题.xlsx")
REPORT_PATH = Path(__file__).resolve().parents[3] / "requirements" / "测试报告.md"
CACHE_PATH = Path(__file__).resolve().parents[3] / "requirements" / "测试结果缓存.json"

SUBJECT_KEYWORDS = [
    "税收收入", "非税收入", "卫生健康支出", "教育支出",
    "社会保障和就业支出", "农林水支出", "科学技术支出",
    "住房保障支出", "交通运输支出", "一般公共预算收入",
    "政府性基金收入", "国有资本经营预算收入", "社会保险基金收入",
]

METRIC_ALIASES = {
    "预算执行率": ["预算执行率", "执行率", "预算完成率"],
    "本月金额": ["本月金额", "当月金额", "本月数", "本月执行金额", "本月执行数"],
    "累计金额": ["累计金额", "累计收入", "累计支出", "累计执行金额", "累计数", "累计执行数"],
    "同比增幅": ["同比增额", "同比增长率", "同比增长", "同比"],
    "环比增幅": ["环比增额", "环比增长率", "环比增长", "环比"],
    "预算数": ["预算数", "年初预算", "调整预算"],
    "总计": ["总计", "合计", "总计多少", "合计多少"],
    "金额": ["金额"],
}


def load_questions() -> List[dict]:
    """从 Excel 读取测试问题并去重。"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["测试用例"]
    seen = set()
    questions = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[2]:
            continue
        question_text = str(row[2]).strip()
        if not question_text or question_text == "问题":
            continue
        key = question_text.rstrip("?？？。！!")
        if key in seen:
            continue
        seen.add(key)
        questions.append({
            "index": len(questions) + 1,
            "business_domain": str(row[0] or "").strip(),
            "question": question_text,
        })
    return questions


def run_tests():
    """批量执行测试。"""
    questions = load_questions()
    print(f"去重后 {len(questions)} 个问题")

    elapsed = 0
    if CACHE_PATH.exists():
        print("发现缓存，跳过 LLM 调用")
        results = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    else:
        llm_client = LlmGateway(LLM_CONFIG)
        extractor = UnifiedIntentExtractor(
            llm_client=llm_client,
            subject_keywords=SUBJECT_KEYWORDS,
            metric_aliases=METRIC_ALIASES,
        )
        dispatcher = ConcurrentTaskDispatcher()

        results = []
        start_time = time.time()

        for i, q in enumerate(questions):
            print(f"[{i+1}/{len(questions)}] {q['question'][:60]}...")
            try:
                d = extractor.extract(q["question"])
                plan = dispatcher.dispatch(d)
                results.append({
                    "idx": q["index"],
                    "question": q["question"],
                    "domain": q["business_domain"],
                    "route": _describe_route(plan),
                    "status": "ok",
                    "t2s_business_module": d.text2sql.business_module,
                    "t2s_account_book": d.text2sql.account_book,
                    "t2s_flow_type": d.text2sql.flow_type,
                    "t2s_region_level": d.text2sql.region_level,
                    "t2s_time_text": d.text2sql.time_text,
                    "t2s_time_start": d.text2sql.time_start,
                    "t2s_time_end": d.text2sql.time_end,
                    "t2s_time_grain": d.text2sql.time_grain,
                    "t2s_query_type": d.text2sql.query_type,
                    "t2s_subjects": d.text2sql.subjects,
                    "t2s_metrics": d.text2sql.metrics,
                    "t2s_regions": d.text2sql.regions,
                    "t2s_data_stage": d.text2sql.data_stage,
                    "t2s_compare_dimension": d.text2sql.compare_dimension,
                    "t2s_compare_operator": d.text2sql.compare_operator,
                    "t2s_chart_hint": d.text2sql.chart_hint,
                    "t2s_top_n": d.text2sql.top_n,
                    "rag_need_policy_basis": d.rag.need_policy_basis,
                    "rag_need_caliber": d.rag.need_caliber_explanation,
                    "rag_need_composition": d.rag.need_composition,
                    "rag_need_data_value": d.rag.need_data_value,
                    "rag_original_question": d.rag.original_question,
                    "dispatch_text2sql": plan.text2sql,
                    "dispatch_rag": plan.rag,
                    "dispatch_clarify": plan.clarify,
                    "dispatch_fallback": plan.fallback,
                    "missing_slots": plan.missing_slots,
                    "clarify_msg": plan.clarify_message[:150] if plan.clarify else "",
                })
            except Exception as e:
                results.append({
                    "idx": q["index"], "question": q["question"],
                    "domain": q["business_domain"], "route": "error",
                    "status": "error", "error": str(e)[:200],
                })

        elapsed = time.time() - start_time
        CACHE_PATH.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")

    _generate_report(results, elapsed, len(questions))
    print(f"报告已生成: {REPORT_PATH}")


def _describe_route(plan) -> str:
    if plan.fallback: return "闲聊兜底"
    if plan.clarify: return "追问补槽位"
    if plan.text2sql and plan.rag: return "混合（Text2SQL + RAG）"
    if plan.text2sql: return "纯 Text2SQL"
    if plan.rag: return "纯 RAG"
    return "未知"


def _generate_report(results: list, elapsed: float, total_questions: int):
    ok = [r for r in results if r["status"] == "ok"]
    err = [r for r in results if r["status"] == "error"]
    t2s_only = [r for r in ok if r["route"] == "纯 Text2SQL"]
    rag_only = [r for r in ok if r["route"] == "纯 RAG"]
    hybrid = [r for r in ok if r["route"] == "混合（Text2SQL + RAG）"]
    clarify = [r for r in ok if r["route"] == "追问补槽位"]
    fallback = [r for r in ok if r["route"] == "闲聊兜底"]

    L = []
    L.append("# 统一意图识别模块 — 测试报告")
    L.append("")
    L.append(f"**测试时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    L.append(f"**数据来源**: 问题.xlsx → 测试用例工作表（已去重）")
    L.append(f"**去重后问题数**: {total_questions}")
    L.append(f"**总耗时**: {elapsed:.1f} 秒")
    L.append("")
    L.append("## 1. 概要")
    L.append("")
    L.append("| 指标 | 数量 |")
    L.append("|------|------|")
    L.append(f"| 去重后问题数 | {total_questions} |")
    L.append(f"| 成功提取 | {len(ok)} |")
    L.append(f"| 提取异常 | {len(err)} |")
    L.append(f"| 成功率 | {len(ok)/len(results)*100:.1f}% |")
    L.append("")
    L.append("## 2. 调度结果分布")
    L.append("")
    L.append("| 调度结果 | 数量 | 占比 |")
    L.append("|----------|------|------|")
    L.append(f"| 纯 Text2SQL | {len(t2s_only)} | {len(t2s_only)/len(ok)*100:.1f}% |")
    L.append(f"| 纯 RAG | {len(rag_only)} | {len(rag_only)/len(ok)*100:.1f}% |")
    L.append(f"| 混合（Text2SQL + RAG） | {len(hybrid)} | {len(hybrid)/len(ok)*100:.1f}% |")
    L.append(f"| 追问补槽位 | {len(clarify)} | {len(clarify)/len(ok)*100:.1f}% |")
    L.append(f"| 闲聊兜底 | {len(fallback)} | {len(fallback)/len(ok)*100:.1f}% |")
    L.append("")

    slot_counts: Dict[str, int] = {}
    for r in clarify:
        for s in r.get("missing_slots", []):
            slot_counts[s] = slot_counts.get(s, 0) + 1
    if slot_counts:
        L.append("### 追问补槽位 — 缺失分析")
        L.append("")
        L.append("| 缺失槽位 | 次数 |")
        L.append("|----------|------|")
        for s, c in sorted(slot_counts.items(), key=lambda x: -x[1]):
            L.append(f"| {s} | {c} |")
        L.append("")

    L.append("## 3. 详细输出（每条问题的完整输入输出）")
    L.append("")

    for r in ok:
        L.append(f"### #{r['idx']} — {r['route']}")
        L.append("")
        L.append(f"**问题**: {r['question']}")
        L.append(f"**Excel业务域**: {r.get('domain','')}")
        L.append("")
        L.append("**text2sql 子集**:")
        L.append("")
        L.append("| 字段 | 值 |")
        L.append("|------|----|")
        for field, key in [
            ("business_module", "t2s_business_module"), ("account_book", "t2s_account_book"),
            ("flow_type", "t2s_flow_type"), ("region_level", "t2s_region_level"),
            ("time_text", "t2s_time_text"), ("time_start", "t2s_time_start"),
            ("time_end", "t2s_time_end"), ("time_grain", "t2s_time_grain"),
            ("query_type", "t2s_query_type"), ("data_stage", "t2s_data_stage"),
            ("compare_dimension", "t2s_compare_dimension"), ("compare_operator", "t2s_compare_operator"),
            ("chart_hint", "t2s_chart_hint"), ("top_n", "t2s_top_n"),
        ]:
            val = r.get(key, "")
            if isinstance(val, list):
                val = ", ".join(val) if val else "—"
            elif not val and val != 0:
                val = "—"
            L.append(f"| {field} | {val} |")
        subjects_val = ", ".join(r.get("t2s_subjects", [])) if r.get("t2s_subjects") else "—"
        metrics_val = ", ".join(r.get("t2s_metrics", [])) if r.get("t2s_metrics") else "—"
        regions_val = ", ".join(r.get("t2s_regions", [])) if r.get("t2s_regions") else "—"
        L.append(f"| subjects | {subjects_val} |")
        L.append(f"| metrics | {metrics_val} |")
        L.append(f"| regions | {regions_val} |")
        L.append("")
        L.append("**rag 子集**:")
        L.append("")
        L.append("| 字段 | 值 |")
        L.append("|------|----|")
        L.append(f"| need_policy_basis | {r.get('rag_need_policy_basis',False)} |")
        L.append(f"| need_caliber_explanation | {r.get('rag_need_caliber',False)} |")
        L.append(f"| need_composition | {r.get('rag_need_composition',False)} |")
        L.append(f"| need_data_value | {r.get('rag_need_data_value',False)} |")
        L.append(f"| original_question | {r.get('rag_original_question','') or '—'} |")
        L.append("")
        L.append("**调度结果**:")
        L.append("")
        L.append(f"- route: **{r['route']}**")
        L.append(f"- text2sql: {r.get('dispatch_text2sql',False)}")
        L.append(f"- rag: {r.get('dispatch_rag',False)}")
        L.append(f"- clarify: {r.get('dispatch_clarify',False)}")
        L.append(f"- fallback: {r.get('dispatch_fallback',False)}")
        if r.get("missing_slots"):
            L.append(f"- missing_slots: {', '.join(r['missing_slots'])}")
        L.append("")
        L.append("---")
        L.append("")

    if err:
        L.append("## 4. 异常")
        L.append("")
        for r in err:
            L.append(f"- **#{r['idx']}** {r['question'][:80]} → {r.get('error','')}")
        L.append("")

    L.append(f"*报告自动生成于 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")
    REPORT_PATH.write_text("\n".join(L), encoding="utf-8")


if __name__ == "__main__":
    run_tests()
