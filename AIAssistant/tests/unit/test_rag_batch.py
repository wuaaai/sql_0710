"""RAG 知识库测试用例批量测试。

读取 知识库测试结果.xlsx 的"08服务器rag"工作表，
对每条 RAG 问题执行 extract → dispatch，生成测试报告。
"""

import json, sys, time
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import openpyxl

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from src.pages.infrastructure.llm.llm_gateway import LlmConfig, LlmGateway
from src.pages.request_lifecycle.dispatch_execution import ConcurrentTaskDispatcher
from src.pages.request_lifecycle.unified_intent_extractor import UnifiedIntentExtractor

LLM_CONFIG = LlmConfig(
    api_key="sk-d507bd835e174d99b57757f3010dfd02",
    base_url="https://api.deepseek.com/v1",
    model="deepseek-chat",
    timeout_seconds=90,
)

EXCEL_PATH = Path(r"E:\Develop_docu\sql_0710\知识库测试结果.xlsx")
REPORT_PATH = Path(__file__).resolve().parents[3] / "requirements" / "测试报告_RAG.md"

SUBJECT_KEYWORDS = [
    "税收收入", "非税收入", "卫生健康支出", "教育支出",
    "社会保障和就业支出", "农林水支出", "科学技术支出",
    "住房保障支出", "交通运输支出", "一般公共预算收入",
    "政府性基金收入", "国有资本经营预算收入", "社会保险基金收入",
]

METRIC_ALIASES = {
    "预算执行率": ["预算执行率", "执行率", "预算完成率"],
    "本月金额": ["本月金额", "当月金额", "本月数", "本月执行金额", "本月执行数"],
    "累计金额": ["累计金额", "累计收入", "累计支出", "累计执行金额", "累计数", "累计执行数"],
    "同比增幅": ["同比增额", "同比增长率", "同比增长", "同比"],
    "环比增幅": ["环比增额", "环比增长率", "环比增长", "环比"],
    "预算数": ["预算数", "年初预算", "调整预算"],
    "总计": ["总计", "合计", "总计多少", "合计多少"],
    "金额": ["金额"],
}


def load_questions() -> List[dict]:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["08服务器rag"]
    questions = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        q = str(row[1]).strip() if row[1] else ""
        if not q or q == "问题":
            continue
        questions.append({"index": len(questions) + 1, "question": q})
    return questions


def run_tests():
    questions = load_questions()
    print(f"加载 {len(questions)} 个 RAG 测试问题")

    llm_client = LlmGateway(LLM_CONFIG)
    extractor = UnifiedIntentExtractor(llm_client, SUBJECT_KEYWORDS, METRIC_ALIASES)
    dispatcher = ConcurrentTaskDispatcher()

    results = []
    start = time.time()
    for i, q in enumerate(questions):
        print(f"[{i+1}/{len(questions)}] {q['question'][:60]}...")
        try:
            d = extractor.extract(q["question"])
            plan = dispatcher.dispatch(d)
            results.append({
                "idx": q["index"], "question": q["question"],
                "route": _describe_route(plan), "status": "ok",
                "t2s_business_module": d.text2sql.business_module,
                "t2s_account_book": d.text2sql.account_book,
                "t2s_flow_type": d.text2sql.flow_type,
                "t2s_region_level": d.text2sql.region_level,
                "t2s_time_text": d.text2sql.time_text,
                "t2s_time_start": d.text2sql.time_start,
                "t2s_time_end": d.text2sql.time_end,
                "t2s_time_grain": d.text2sql.time_grain,
                "t2s_query_type": d.text2sql.query_type,
                "t2s_subjects": d.text2sql.subjects,
                "t2s_metrics": d.text2sql.metrics,
                "t2s_regions": d.text2sql.regions,
                "t2s_data_stage": d.text2sql.data_stage,
                "t2s_compare_dimension": d.text2sql.compare_dimension,
                "t2s_compare_operator": d.text2sql.compare_operator,
                "t2s_chart_hint": d.text2sql.chart_hint,
                "t2s_top_n": d.text2sql.top_n,
                "rag_need_policy_basis": d.rag.need_policy_basis,
                "rag_need_caliber": d.rag.need_caliber_explanation,
                "rag_need_composition": d.rag.need_composition,
                "rag_need_data_value": d.rag.need_data_value,
                "dispatch_text2sql": plan.text2sql,
                "dispatch_rag": plan.rag,
                "dispatch_clarify": plan.clarify,
                "dispatch_fallback": plan.fallback,
                "missing_slots": plan.missing_slots,
            })
        except Exception as e:
            results.append({"idx": q["index"], "question": q["question"], "status": "error", "error": str(e)[:200]})

    elapsed = time.time() - start
    _generate_report(results, elapsed, len(questions))
    print(f"报告: {REPORT_PATH}")


def _describe_route(plan) -> str:
    if plan.fallback: return "闲聊兜底"
    if plan.clarify: return "追问补槽位"
    if plan.text2sql and plan.rag: return "混合（Text2SQL + RAG）"
    if plan.text2sql: return "纯 Text2SQL"
    if plan.rag: return "纯 RAG"
    return "未知"


def _generate_report(results, elapsed, total):
    ok = [r for r in results if r["status"] == "ok"]
    t2s_only = [r for r in ok if r["route"] == "纯 Text2SQL"]
    rag_only = [r for r in ok if r["route"] == "纯 RAG"]
    hybrid = [r for r in ok if r["route"] == "混合（Text2SQL + RAG）"]
    clarify = [r for r in ok if r["route"] == "追问补槽位"]

    L = []
    L.append("# 统一意图识别模块 — RAG 知识库测试报告")
    L.append("")
    L.append(f"**测试时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    L.append(f"**数据来源**: 知识库测试结果.xlsx → 08服务器rag")
    L.append(f"**问题数**: {total}")
    L.append(f"**总耗时**: {elapsed:.1f} 秒")
    L.append("")
    L.append("## 1. 概要")
    L.append("")
    L.append(f"| 总问题数 | {total} |")
    L.append(f"| 成功提取 | {len(ok)} |")
    L.append(f"| 成功率 | {len(ok)/len(results)*100:.1f}% |")
    L.append("")
    L.append("## 2. 调度结果分布")
    L.append("")
    L.append("| 调度结果 | 数量 | 占比 |")
    L.append("|----------|------|------|")
    L.append(f"| 纯 Text2SQL | {len(t2s_only)} | {len(t2s_only)/len(ok)*100:.1f}% |")
    L.append(f"| 纯 RAG | {len(rag_only)} | {len(rag_only)/len(ok)*100:.1f}% |")
    L.append(f"| 混合（Text2SQL + RAG） | {len(hybrid)} | {len(hybrid)/len(ok)*100:.1f}% |")
    L.append(f"| 追问补槽位 | {len(clarify)} | {len(clarify)/len(ok)*100:.1f}% |")
    L.append("")

    L.append("## 3. 详细输出")
    L.append("")
    for r in ok:
        L.append(f"### #{r['idx']} — {r['route']}")
        L.append(f"**问题**: {r['question']}")
        L.append("")
        L.append("**text2sql 子集**:")
        L.append("| 字段 | 值 |")
        L.append("|------|----|")
        for f, k in [("business_module","t2s_business_module"),("account_book","t2s_account_book"),
                     ("flow_type","t2s_flow_type"),("region_level","t2s_region_level"),
                     ("time_text","t2s_time_text"),("time_start","t2s_time_start"),
                     ("time_end","t2s_time_end"),("time_grain","t2s_time_grain"),
                     ("query_type","t2s_query_type"),("data_stage","t2s_data_stage"),
                     ("compare_dimension","t2s_compare_dimension"),("compare_operator","t2s_compare_operator"),
                     ("chart_hint","t2s_chart_hint"),("top_n","t2s_top_n")]:
            v = r.get(k, "")
            if isinstance(v, list): v = ", ".join(v) if v else "—"
            elif not v and v != 0: v = "—"
            L.append(f"| {f} | {v} |")
        for label, key in [("subjects","t2s_subjects"),("metrics","t2s_metrics"),("regions","t2s_regions")]:
            v = r.get(key, [])
            L.append(f"| {label} | {', '.join(v) if v else '—'} |")
        L.append("")
        L.append("**rag 子集**:")
        L.append("| 字段 | 值 |")
        L.append("|------|----|")
        L.append(f"| need_policy_basis | {r.get('rag_need_policy_basis',False)} |")
        L.append(f"| need_caliber_explanation | {r.get('rag_need_caliber',False)} |")
        L.append(f"| need_composition | {r.get('rag_need_composition',False)} |")
        L.append(f"| need_data_value | {r.get('rag_need_data_value',False)} |")
        L.append("")
        L.append(f"**调度**: {r['route']} | text2sql={r.get('dispatch_text2sql')} | rag={r.get('dispatch_rag')}")
        if r.get("missing_slots"):
            L.append(f" | 缺失: {', '.join(r['missing_slots'])}")
        L.append("")
        L.append("---")
        L.append("")

    L.append(f"*报告生成于 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")
    REPORT_PATH.write_text("\n".join(L), encoding="utf-8")


if __name__ == "__main__":
    run_tests()
