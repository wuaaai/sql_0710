"""RAG 反哺 Text2SQL 模块 — 真实数据效果测试。

读取两套测试集，对每条问题执行完整管线：提取 → RAG检索 → 反哺增强 → 调度，
对比 v1/v2 意图差异，生成效果报告到 requirements/反哺效果报告.md。
"""

import json, sys, time
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import openpyxl

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from src.configs.datasource_settings import DataSourceSettings
from src.pages.infrastructure.llm.llm_gateway import LlmConfig, LlmGateway
from src.pages.infrastructure.vectorstores.vector_store_client import PgVectorStore
from src.pages.rag_knowledge_chain.retrieval import CustomLocalEmbeddings, HybridRetriever
from src.pages.request_lifecycle.unified_intent_extractor import UnifiedIntentExtractor
from src.pages.request_lifecycle.rag_slot_enricher import RagSlotEnricher
from src.pages.request_lifecycle.dispatch_execution import ConcurrentTaskDispatcher
from src.pages.request_lifecycle.rag_driven_text2sql_pipeline import RagDrivenText2SqlPipeline
from src.services.rag_service import RagService

# ============================================================
# 配置
# ============================================================
ds = DataSourceSettings()

LLM_CONFIG = LlmConfig(
    api_key="sk-fdc7d21b49ca45c08b5962fee5f4847f",
    base_url="https://api.deepseek.com/v1",
    model="deepseek-chat", timeout_seconds=90,
)

SUBJECT_KEYWORDS = [
    "税收收入", "非税收入", "卫生健康支出", "教育支出",
    "社会保障和就业支出", "农林水支出", "科学技术支出",
    "住房保障支出", "交通运输支出", "一般公共预算收入",
    "政府性基金收入", "国有资本经营预算收入", "社会保险基金收入",
]

METRIC_ALIASES = {
    "预算执行率": ["预算执行率", "执行率", "预算完成率"],
    "本月金额": ["本月金额", "当月金额", "本月数", "本月执行金额", "本月执行数"],
    "累计金额": ["累计金额", "累计收入", "累计支出", "累计执行金额", "累计数", "累计执行数"],
    "同比增幅": ["同比增额", "同比增长率", "同比增长", "同比"],
    "环比增幅": ["环比增额", "环比增长率", "环比增长", "环比"],
    "预算数": ["预算数", "年初预算", "调整预算"],
    "总计": ["总计", "合计", "总计多少", "合计多少"],
    "金额": ["金额"],
}

REPORT_PATH = Path(__file__).resolve().parents[3] / "requirements" / "反哺效果报告.md"


def load_questions(path, sheet):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    seen = set()
    qs = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if sheet == "08服务器rag":
            q = str(row[1]).strip() if row[1] else ""
        else:
            q = str(row[2]).strip() if row[2] else ""
        if not q or q == "问题":
            continue
        key = q.rstrip("?？？。！!")
        if key in seen:
            continue
        seen.add(key)
        qs.append({"question": q, "index": len(qs) + 1})
    return qs


def setup_services():
    llm = LlmGateway(LLM_CONFIG)
    extractor = UnifiedIntentExtractor(llm, SUBJECT_KEYWORDS, METRIC_ALIASES)
    enricher = RagSlotEnricher(METRIC_ALIASES, SUBJECT_KEYWORDS)
    dispatcher = ConcurrentTaskDispatcher()

    rag_cfg = ds.rag_vector
    emb = CustomLocalEmbeddings(api_url="http://10.32.10.160:8991/embed")
    vector_store = PgVectorStore(connection=rag_cfg.connection_string, collection_name=rag_cfg.collection_name, embedding=emb)
    retriever = HybridRetriever(vector_store=vector_store, rerank_url="http://10.32.10.160:8991/rerank", top_k=5, recall_k=10)
    rag_service = RagService(retriever=retriever)

    return extractor, enricher, dispatcher, rag_service


def run():
    t2s_qs = load_questions(r"E:\Develop_docu\sql_0710\问题.xlsx", "测试用例")
    rag_qs = load_questions(r"E:\Develop_docu\sql_0710\知识库测试结果.xlsx", "08服务器rag")
    all_qs = t2s_qs + rag_qs
    # 跨文件去重
    seen = set()
    unique = []
    for q in all_qs:
        key = q["question"].rstrip("?？？。！!")
        if key not in seen:
            seen.add(key)
            q["index"] = len(unique) + 1
            unique.append(q)

    print(f"text2sql测试集: {len(t2s_qs)} 条, RAG测试集: {len(rag_qs)} 条, 去重后: {len(unique)} 条")

    extractor, enricher, dispatcher, rag_service = setup_services()
    results = []
    start = time.time()

    for i, q in enumerate(unique):
        print(f"[{i+1}/{len(unique)}] {q['question'][:60]}...")
        try:
            # 1. 提取 v1
            intent_v1 = extractor.extract(q["question"])
            t2s_v1 = intent_v1.text2sql
            v1_ready = _slots_ready_static(t2s_v1)

            # 2. RAG 检索
            rag_result = rag_service.search(intent_v1.rag)
            snippet_count = len(rag_result.snippets) if rag_result.snippets else 0
            top_score = max((s.score for s in rag_result.snippets), default=0) if rag_result.snippets else 0

            # 3. 反哺增强
            enriched = False
            filled_slots = []
            if t2s_v1.subjects or t2s_v1.metrics:
                intent_v2, log = enricher.enrich(intent_v1, rag_result.snippets, q["question"])
                filled_slots = log.filled_slots
                enriched = len(filled_slots) > 0
                v2_ready = _slots_ready_static(intent_v2.text2sql) if enriched else v1_ready
                intent_final = intent_v2 if enriched else intent_v1
            else:
                intent_final = intent_v1
                v2_ready = v1_ready

            # 4. 调度
            plan = dispatcher.dispatch(intent_final)

            results.append({
                "idx": q["index"], "question": q["question"],
                "v1_slots_ready": v1_ready,
                "v1_region": t2s_v1.region_level, "v1_metrics": t2s_v1.metrics,
                "v1_time": t2s_v1.time_text, "v1_account": t2s_v1.account_book,
                "v1_data_stage": t2s_v1.data_stage, "v1_subjects": t2s_v1.subjects,
                "rag_snippets": snippet_count, "rag_top_score": round(top_score, 3),
                "enriched": enriched, "filled_slots": filled_slots,
                "v2_slots_ready": v2_ready,
                "route": _descr(plan),
                "status": "ok",
            })
        except Exception as e:
            results.append({"idx": q["index"], "question": q["question"], "status": "error", "error": str(e)[:200]})

    elapsed = time.time() - start
    _gen_report(results, elapsed, len(unique), len(t2s_qs), len(rag_qs))
    print(f"报告: {REPORT_PATH}")


def _slots_ready_static(t2s) -> bool:
    return bool(t2s.subjects and t2s.metrics and t2s.flow_type in ("收入", "支出") and t2s.region_level)


def _descr(plan) -> str:
    if plan.fallback: return "fallback"
    if plan.clarify: return "clarify"
    if plan.text2sql and plan.rag: return "both"
    if plan.text2sql: return "sql"
    if plan.rag: return "rag"
    return "?"


def _gen_report(results, elapsed, total, t2s_cnt, rag_cnt):
    ok = [r for r in results if r["status"] == "ok"]
    enriched_list = [r for r in ok if r["enriched"]]
    v1_ready = [r for r in ok if r["v1_slots_ready"]]
    v2_ready = [r for r in ok if r["v2_slots_ready"]]
    newly_ready = [r for r in ok if not r["v1_slots_ready"] and r["v2_slots_ready"]]

    L = []
    L.append("# RAG 反哺 Text2SQL — 效果测试报告")
    L.append("")
    L.append(f"**测试时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    L.append(f"**数据来源**: 问题.xlsx({t2s_cnt}条) + 知识库测试结果.xlsx({rag_cnt}条) → 去重后 {total} 条")
    L.append(f"**总耗时**: {elapsed:.1f} 秒")
    L.append("")
    L.append("## 1. 概要")
    L.append("")
    L.append("| 指标 | 数量 | 占比 |")
    L.append("|------|------|------|")
    L.append(f"| 总问题数 | {total} | 100% |")
    L.append(f"| 成功执行 | {len(ok)} | {len(ok)/len(results)*100:.1f}% |")
    L.append(f"| v1 槽位齐全 | {len(v1_ready)} | {len(v1_ready)/len(ok)*100:.1f}% |")
    L.append(f"| RAG 反哺触发 | {len(enriched_list)} | {len(enriched_list)/len(ok)*100:.1f}% |")
    L.append(f"| **反哺后新增齐全** | **{len(newly_ready)}** | **{len(newly_ready)/len(ok)*100:.1f}%** |")
    L.append(f"| v2 槽位齐全 | {len(v2_ready)} | {len(v2_ready)/len(ok)*100:.1f}% |")
    L.append("")
    L.append("## 2. 调度分布")
    L.append("")
    routes = {}
    for r in ok:
        routes[r["route"]] = routes.get(r["route"], 0) + 1
    L.append("| 调度结果 | 数量 | 占比 |")
    L.append("|----------|------|------|")
    for rt in ["sql", "rag", "both", "clarify", "fallback"]:
        if rt in routes:
            L.append(f"| {rt} | {routes[rt]} | {routes[rt]/len(ok)*100:.1f}% |")
    L.append("")

    # 反哺新增齐全的问题
    if newly_ready:
        L.append("## 3. 反哺新增槽位齐全的问题（关键效果）")
        L.append("")
        L.append(f"共 {len(newly_ready)} 条问题通过 RAG 反哺补齐了缺失槽位，从 clarify 变为可执行 sql/both：")
        L.append("")
        L.append("| # | 问题 | 补全槽位 | v1缺失 |")
        L.append("|---|------|---------|--------|")
        for r in newly_ready:
            missing_v1 = []
            if not r.get("v1_region"): missing_v1.append("region")
            if not r.get("v1_metrics"): missing_v1.append("metrics")
            if not r.get("v1_time"): missing_v1.append("time")
            if not r.get("v1_account"): missing_v1.append("account")
            L.append(f"| {r['idx']} | {r['question'][:50]} | {', '.join(r['filled_slots'])} | {', '.join(missing_v1)} |")
        L.append("")

    # 反哺统计
    if enriched_list:
        L.append("## 4. 反哺槽位统计")
        L.append("")
        slot_count = {}
        for r in enriched_list:
            for s in r["filled_slots"]:
                slot_count[s] = slot_count.get(s, 0) + 1
        L.append("| 槽位 | 被补全次数 |")
        L.append("|------|----------|")
        for s, c in sorted(slot_count.items(), key=lambda x: -x[1]):
            L.append(f"| {s} | {c} |")
        L.append("")

    # 全部详细结果
    L.append("## 5. 全部问题详情")
    L.append("")
    L.append("| # | 问题(前40字) | v1齐全 | 反哺 | 补全槽位 | v2齐全 | 调度 | RAG分 |")
    L.append("|---|-------------|:----:|:---:|---------|:----:|:----:|:-----:|")
    for r in ok:
        enriched_str = "Y" if r["enriched"] else "—"
        filled_str = ", ".join(r["filled_slots"]) if r["filled_slots"] else "—"
        L.append(f"| {r['idx']} | {r['question'][:40]} | {'Y' if r['v1_slots_ready'] else '—'} | {enriched_str} | {filled_str} | {'Y' if r['v2_slots_ready'] else '—'} | {r['route']} | {r['rag_top_score']} |")
    L.append("")

    L.append("## 6. 补充反哺详细数据")
    L.append("")
    for r in enriched_list[:20]:  # 前20条反哺详情
        L.append(f"### #{r['idx']} {r['question'][:60]}")
        L.append(f"- v1: region={r.get('v1_region','')}, metrics={r.get('v1_metrics','')}, time={r.get('v1_time','')}, account={r.get('v1_account','')}, data_stage={r.get('v1_data_stage','')}")
        L.append(f"- 补全: {r['filled_slots']}")
        L.append(f"- RAG: {r['rag_snippets']} snippets, top_score={r['rag_top_score']}")
        L.append(f"- v1齐全→v2齐全: {r['v1_slots_ready']}→{r['v2_slots_ready']}")
        L.append(f"- 调度: {r['route']}")
        L.append("")
    L.append("")

    L.append(f"*报告生成于 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")
    REPORT_PATH.write_text("\n".join(L), encoding="utf-8")


if __name__ == "__main__":
    run()
